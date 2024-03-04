import requests
import re
import datetime
import calendar
import sys
import json
from openpyxl import load_workbook
from flask import request, jsonify
from DefaultInfo import *
from datetime import date, timedelta

URL = URL_SCHOOL_SCHEDULE_INFO

def SchoolExamInfo():
    weekly      = {0:"월", 1:"화", 2:"수", 3:"목", 4:"금", 5:"토", 6:"일"}   
    req = request.get_json() 

    if bool('grade' in req['action']['detailParams']) == True and bool('semester' in req['action']['detailParams']) == True and bool('term' in req['action']['detailParams']) == True:    
        grade = req['action']['detailParams']['grade']['value']
        semester = req['action']['detailParams']['semester']['value']
        term = req['action']['detailParams']['term']['value']
        if grade == '1학년':
            Error = {
                "version": "2.0",
                "data": {
                    'Exam_Info': "입력이 잘못되었습니다.",
                    'Exam_Scope':"-"
                }
            }
            return jsonify(Error)              

        load_wb = load_workbook(PATH + "UserData_writtentest.xlsx", data_only=True)
        load_ws = load_wb[grade + " " + semester + " " + term]

        bAllRender = load_ws['C1'].value
        if bAllRender != 'y':
            Error = {
                "version": "2.0",
                "data": {
                    'Exam_Info': "시험 기간이 아니에요.",
                    'Exam_Scope':""
                }
            }
            return jsonify(Error)           

        # max_row 
        max_row = load_ws.max_row
        
        # max_column 
        max_column = load_ws.max_column
 
        examInfo = []   
        # max row, col
        for i in range(3, max_row + 1):
            bRender = load_ws.cell(row=i, column=6).value
            if bRender != 'y' and bRender != 'None':
                continue
            dict_info = {}
            dict_info['date'] = load_ws.cell(row=i, column=1).value
            dict_info['period'] = load_ws.cell(row=i, column=2).value
            dict_info['start_time'] = load_ws.cell(row=i, column=3).value
            dict_info['end_time'] = load_ws.cell(row=i, column=4).value
            dict_info['subject_name'] = load_ws.cell(row=i, column=5).value
            dict_info['subject_info'] = str(load_ws.cell(row=i, column=7).value)
            #print(str(dict_info), file=sys.stderr)
            examInfo.append(dict_info)
        print("지필고사 데이터:")    
        print(str(examInfo), file=sys.stderr)
        examInfoText = "[시험 시간표]\n"
        examScopeText = "[시험 범위]\n"
        preDate = datetime.datetime(1,1,1)
        for i in range(len(examInfo)):
            examInfo[i]['subject_info'] = str(examInfo[i]['subject_info']).replace('%n', '\n')
            examScopeText += "-" + str(examInfo[i]['subject_name']) + " " + str(examInfo[i]['subject_info']) + "\n"

            if preDate != examInfo[i]['date']:
                ListDate = str(examInfo[i]['date'].strftime('%y-%m-%d')).split('-')

                monthValue = str(ListDate[1]).zfill(2)
                dayValue = str(ListDate[2]).zfill(2)
                targetDate = datetime.date(int(ListDate[0]), int(ListDate[1]), int(ListDate[2]))
                examInfoText = examInfoText + monthValue + "월" + " " + dayValue + "일" + " " + "(" + weekly[targetDate.weekday()] + ")" + "\n"
                examInfoText = examInfoText + "-" + str(examInfo[i]['period']) + "교시" + " " + "(" + str(examInfo[i]['start_time'].strftime("%H:%M")) + "~" + str(examInfo[i]['end_time'].strftime("%H:%M")) + ")" + "(" + str(int(examInfo[i]['end_time'].hour * 60 + examInfo[i]['end_time'].minute) -  int(examInfo[i]['start_time'].hour * 60 + examInfo[i]['start_time'].minute)) + "분" + ")" + " " + str(examInfo[i]['subject_name']) + "\n"
                preDate = examInfo[i]['date']
            else:
               examInfoText = examInfoText + "-" + str(examInfo[i]['period']) + "교시" + " " + "(" + str(examInfo[i]['start_time'].strftime("%H:%M")) + "~" + str(examInfo[i]['end_time'].strftime("%H:%M")) + ")" + "(" + str(int(examInfo[i]['end_time'].hour * 60 + examInfo[i]['end_time'].minute) -  int(examInfo[i]['start_time'].hour * 60 + examInfo[i]['start_time'].minute)) + "분" + ")" + " " + str(examInfo[i]['subject_name']) + "\n"

        print(str(examInfoText), file=sys.stderr)
        responseData = {
            "version": "2.0",
            "data": {
                'Exam_Info': examInfoText,
                'Exam_Scope' : examScopeText
            }
        }
        return jsonify(responseData)

    Error = {
                "version": "2.0",
                "data": {
                    'Exam_Info': "입력이 잘못되었습니다.",
                    'Exam_Scope':"-"
                }
            }
            

    return jsonify(Error) 
    