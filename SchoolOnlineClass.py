import requests
import re
import datetime
import calendar
import sys
import json
from openpyxl import load_workbook
from flask import request, jsonify
from DefaultInfo import *
from datetime import date, timedelta

def SchoolOnlineClass():
    req = request.get_json() 

    if bool('grade' in req['action']['detailParams']) == True:    

        if bool('sys_date_period' in req['action']['detailParams']) == True:
            print("변환 하기 전 데이터: ")    
            print(str(req['action']['detailParams']['sys_date_period']['value']), file=sys.stderr)
            loadedJson =  json.loads(req['action']['detailParams']['sys_date_period']['value'])
            fromDate = str(loadedJson['from']['date'])
            toDate = str(loadedJson['to']['date'])
            fromList = fromDate.split('-')
            toList = toDate.split('-')
            fromDate = datetime.datetime(int(fromList[0]), int(fromList[1]), int(fromList[2]), 0, 0) 
            toDate = datetime.datetime(int(toList[0]), int(toList[1]), int(toList[2]), 0, 0) 
            requestParameter_FROM = fromDate
            requestParameter_TO = toDate
            print("변환 후 데이터: ")
            print(requestParameter_FROM)
            print(requestParameter_TO)   

        elif bool('sys_date' in req['action']['detailParams']) == True:
            loadedJson =  json.loads(req['action']['detailParams']['sys_date']['value'])
            targetDate = str(loadedJson['date'])
            
            if 'None' == targetDate:
                Error = {
                    "version": "2.0",
                    "data": {
                        'OnlineClass_Info': "입력이 잘못되었습니다."
                    }
                }
                return jsonify(Error) 
            
            targetList = targetDate.split('-')
            
            targetDate = datetime.datetime(int(targetList[0]), int(targetList[1]), int(targetList[2]), 0, 0) 
            requestParameter_FROM = targetDate
            requestParameter_TO = targetDate
        else:
            Error = {
                "version": "2.0",
                "data": {
                    'OnlineClass_Info': "입력이 잘못되었습니다."
                }
            }
            return jsonify(Error)             

        grade = req['action']['detailParams']['grade']['value']
        grade = re.sub(r'[^0-9]', '', grade)
        load_wb = load_workbook(PATH + "UserData_onoffclass.xlsx", data_only=True)
        load_ws = load_wb["데이터시트"]
        
        
        # max_row 
        max_row = load_ws.max_row
    
        OnlineInfo = {}  
        for i in range(2, max_row + 1):
            dict_info = {}
            dict_info['date'] = load_ws.cell(row=i, column=1).value
            dict_info['week'] = load_ws.cell(row=i, column=2).value
            dict_info['1'] = load_ws.cell(row=i, column=3).value
            dict_info['2'] = load_ws.cell(row=i, column=4).value
            dict_info['3'] = load_ws.cell(row=i, column=5).value
            idxList = str(dict_info['date'].strftime('%y-%m-%d')).split('-')
            idx = idxList[0] + idxList[1] + idxList[2]
            OnlineInfo[str(load_ws.cell(row=i, column=1).value)] = dict_info
        print("엑셀 데이터: ")
        print(str(OnlineInfo), file=sys.stderr)


        InfoText = ""

        for i in OnlineInfo:   
     
            if str(requestParameter_FROM) == i: 
                ListDate = str(OnlineInfo[i]['date'].strftime('%y-%m-%d')).split('-')
                monthValue = str(ListDate[1]).zfill(2)
                dayValue = str(ListDate[2]).zfill(2)
                targetDate = datetime.date(int(ListDate[0]), int(ListDate[1]), int(ListDate[2]))
                InfoText = InfoText + monthValue + "월" + " " + dayValue + "일" + " " + "(" + OnlineInfo[i]['week'] + ")" + " " 
                InfoText = InfoText + OnlineInfo[i][str(grade)] + "\n"
                requestParameter_FROM = requestParameter_FROM + timedelta(days=1)
            if str(requestParameter_TO) == i:
                break

        responseData = {
            "version": "2.0",
            "data": {
                'OnlineClass_Info': InfoText
            }
        }
        return jsonify(responseData)

    Error = {
                "version": "2.0",
                "data": {
                    'OnlineClass_Info': "입력이 잘못되었습니다."
                }
            }
            

    return jsonify(Error) 